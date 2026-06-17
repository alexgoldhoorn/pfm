"""
Portfolio Dividend Tracker (PDT) XLSX import/export.

Handles the Google Sheets template format from app.portfoliodividendtracker.com.
The spreadsheet has three data sheets:
  - Transactions: buy/sell orders with costs and taxes
  - Dividends:    dividend payments
  - Bookings:     deposits and withdrawals

Each sheet uses a 3-row header block (machine keys, group labels, display labels)
followed by data rows starting at row 4.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any, Iterator, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover
    raise ImportError(
        "openpyxl is required for PDT XLSX support. "
        "Install it with: pip install openpyxl"
    )


# ---------------------------------------------------------------------------
# Data containers
# ---------------------------------------------------------------------------


@dataclass
class PDTTransaction:
    """A single buy/sell row from the PDT Transactions sheet."""

    broker: str
    name: str
    pdt_type: str  # 'Stock market', 'Crypto', 'Commodity'
    search: str  # ISIN or ticker
    exchange: str
    date: date
    action: str  # 'Buy' or 'Sell'
    amount: float  # quantity
    price: float
    price_currency: str
    price_exchange_rate: Optional[float] = None
    price_exchange_rate_currency: Optional[str] = None
    costs: Optional[float] = None
    costs_currency: Optional[str] = None
    costs_exchange_rate: Optional[float] = None
    costs_exchange_rate_currency: Optional[str] = None
    tax: Optional[float] = None
    tax_currency: Optional[str] = None
    tax_exchange_rate: Optional[float] = None
    tax_exchange_rate_currency: Optional[str] = None


@dataclass
class PDTDividend:
    """A single row from the PDT Dividends sheet."""

    broker: str
    name: str
    pdt_type: str
    search: str
    exchange: str
    date: date
    action: str  # 'Cash', 'Stock', 'Staking'
    amount: float
    amount_currency: str
    amount_exchange_rate: Optional[float] = None
    amount_exchange_rate_currency: Optional[str] = None
    tax: Optional[float] = None
    tax_currency: Optional[str] = None
    tax_exchange_rate: Optional[float] = None
    tax_exchange_rate_currency: Optional[str] = None
    costs: Optional[float] = None
    costs_currency: Optional[str] = None
    costs_exchange_rate: Optional[float] = None
    costs_exchange_rate_currency: Optional[str] = None


@dataclass
class PDTBooking:
    """A single row from the PDT Bookings sheet."""

    broker: str
    date: date
    action: str  # 'Deposit', 'Withdrawal'
    amount: float
    currency: str


@dataclass
class PDTParseResult:
    """Aggregated result of parsing a PDT XLSX file."""

    transactions: List[PDTTransaction] = field(default_factory=list)
    dividends: List[PDTDividend] = field(default_factory=list)
    bookings: List[PDTBooking] = field(default_factory=list)
    skipped: List[Tuple[str, str]] = field(default_factory=list)  # (sheet, reason)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_ETF_KEYWORDS = re.compile(r"\bETF\b|UCITS|INDEX\s+FUND|FUND\b|TRUST\b", re.IGNORECASE)


def _detect_asset_type(name: str, pdt_type: str) -> str:
    """Map a PDT effect type + name to our internal asset_type string."""
    pt = pdt_type.lower()
    if pt == "crypto":
        return "crypto"
    if pt == "commodity":
        return "commodity"
    # 'Stock market' — heuristic on name
    if _ETF_KEYWORDS.search(name):
        return "etf"
    return "stock"


def _pdt_action_to_tx_type(action: str) -> Optional[str]:
    """Convert a PDT transaction-action to our transaction_type."""
    mapping = {"buy": "buy", "sell": "sell"}
    return mapping.get(action.lower())


def _tx_type_to_pdt_action(tx_type: str) -> Optional[str]:
    """Convert our transaction_type to a PDT transaction-action."""
    mapping = {"buy": "Buy", "sell": "Sell"}
    return mapping.get(tx_type.lower())


def _asset_type_to_pdt_type(asset_type: str) -> str:
    """Convert our asset_type to a PDT effect type string."""
    mapping = {
        "stock": "Stock market",
        "etf": "Stock market",
        "index": "Stock market",
        "mutual_fund": "Stock market",
        "bond": "Stock market",
        "crypto": "Crypto",
        "commodity": "Commodity",
        "cash": "Stock market",
    }
    return mapping.get(asset_type.lower(), "Stock market")


def _to_date(value: Any) -> Optional[date]:
    """Coerce a cell value to a date object."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _float_or_none(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _str_or_none(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _iter_data_rows(
    ws: Worksheet, n_header_rows: int = 3
) -> Iterator[Tuple[int, tuple]]:
    """Yield (row_number, values_tuple) for data rows, skipping header rows."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= n_header_rows:
            continue
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        yield row_idx, row


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


class PDTXLSXParser:
    """Parses a Portfolio Dividend Tracker v2 XLSX file into structured data."""

    def parse(self, file_path: str) -> PDTParseResult:
        """
        Parse all relevant sheets from a PDT XLSX file.

        Args:
            file_path: Path to the .xlsx file.

        Returns:
            PDTParseResult containing all parsed rows and any skipped entries.
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        result = PDTParseResult()

        # Prefer the canonical sheet names; fall back to first match by prefix
        tx_sheet = self._get_sheet(wb, "Transactions")
        div_sheet = self._get_sheet(wb, "Dividends")
        book_sheet = self._get_sheet(wb, "Bookings")

        if tx_sheet:
            self._parse_transactions(tx_sheet, result)
        if div_sheet:
            self._parse_dividends(div_sheet, result)
        if book_sheet:
            self._parse_bookings(book_sheet, result)

        return result

    def _get_sheet(self, wb: Workbook, name: str) -> Optional[Worksheet]:
        if name in wb.sheetnames:
            return wb[name]
        return None

    def _parse_transactions(self, ws: Worksheet, result: PDTParseResult) -> None:
        for row_idx, row in _iter_data_rows(ws):
            # Column order (0-indexed):
            # 0:broker 1:name 2:type 3:search 4:exchange 5:date 6:time
            # 7:action 8:amount 9:price 10:price-cur 11:price-xr 12:price-xr-cur
            # 13:costs 14:costs-cur 15:costs-xr 16:costs-xr-cur
            # 17:tax 18:tax-cur 19:tax-xr 20:tax-xr-cur
            try:
                broker = _str_or_none(row[0])
                name = _str_or_none(row[1])
                pdt_type = _str_or_none(row[2]) or "Stock market"
                search = _str_or_none(row[3]) or ""
                exchange = _str_or_none(row[4]) or ""
                tx_date = _to_date(row[5])
                action = _str_or_none(row[7])
                amount = _float_or_none(row[8])
                price = _float_or_none(row[9])
                price_currency = _str_or_none(row[10])

                if (
                    not all([broker, name, tx_date, action, price_currency])
                    or amount is None
                    or price is None
                ):
                    result.skipped.append(
                        ("Transactions", f"Row {row_idx}: missing required fields")
                    )
                    continue

                result.transactions.append(
                    PDTTransaction(
                        broker=broker,
                        name=name,
                        pdt_type=pdt_type,
                        search=search,
                        exchange=exchange,
                        date=tx_date,
                        action=action,
                        amount=amount,
                        price=price,
                        price_currency=price_currency,
                        price_exchange_rate=(
                            _float_or_none(row[11]) if len(row) > 11 else None
                        ),
                        price_exchange_rate_currency=(
                            _str_or_none(row[12]) if len(row) > 12 else None
                        ),
                        costs=_float_or_none(row[13]) if len(row) > 13 else None,
                        costs_currency=_str_or_none(row[14]) if len(row) > 14 else None,
                        costs_exchange_rate=(
                            _float_or_none(row[15]) if len(row) > 15 else None
                        ),
                        costs_exchange_rate_currency=(
                            _str_or_none(row[16]) if len(row) > 16 else None
                        ),
                        tax=_float_or_none(row[17]) if len(row) > 17 else None,
                        tax_currency=_str_or_none(row[18]) if len(row) > 18 else None,
                        tax_exchange_rate=(
                            _float_or_none(row[19]) if len(row) > 19 else None
                        ),
                        tax_exchange_rate_currency=(
                            _str_or_none(row[20]) if len(row) > 20 else None
                        ),
                    )
                )
            except Exception as e:
                result.skipped.append(("Transactions", f"Row {row_idx}: {e}"))

    def _parse_dividends(self, ws: Worksheet, result: PDTParseResult) -> None:
        for row_idx, row in _iter_data_rows(ws):
            # 0:broker 1:name 2:type 3:search 4:exchange 5:date 6:time
            # 7:action 8:amount 9:amount-cur 10:amount-xr 11:amount-xr-cur
            # 12:tax 13:tax-cur 14:tax-xr 15:tax-xr-cur
            # 16:costs 17:costs-cur 18:costs-xr 19:costs-xr-cur
            try:
                broker = _str_or_none(row[0])
                name = _str_or_none(row[1])
                pdt_type = _str_or_none(row[2]) or "Stock market"
                search = _str_or_none(row[3]) or ""
                exchange = _str_or_none(row[4]) or ""
                tx_date = _to_date(row[5])
                action = _str_or_none(row[7])
                amount = _float_or_none(row[8])
                amount_currency = _str_or_none(row[9])

                if (
                    not all([broker, name, tx_date, action, amount_currency])
                    or amount is None
                ):
                    result.skipped.append(
                        ("Dividends", f"Row {row_idx}: missing required fields")
                    )
                    continue

                result.dividends.append(
                    PDTDividend(
                        broker=broker,
                        name=name,
                        pdt_type=pdt_type,
                        search=search,
                        exchange=exchange,
                        date=tx_date,
                        action=action,
                        amount=amount,
                        amount_currency=amount_currency,
                        amount_exchange_rate=(
                            _float_or_none(row[10]) if len(row) > 10 else None
                        ),
                        amount_exchange_rate_currency=(
                            _str_or_none(row[11]) if len(row) > 11 else None
                        ),
                        tax=_float_or_none(row[12]) if len(row) > 12 else None,
                        tax_currency=_str_or_none(row[13]) if len(row) > 13 else None,
                        tax_exchange_rate=(
                            _float_or_none(row[14]) if len(row) > 14 else None
                        ),
                        tax_exchange_rate_currency=(
                            _str_or_none(row[15]) if len(row) > 15 else None
                        ),
                        costs=_float_or_none(row[16]) if len(row) > 16 else None,
                        costs_currency=_str_or_none(row[17]) if len(row) > 17 else None,
                        costs_exchange_rate=(
                            _float_or_none(row[18]) if len(row) > 18 else None
                        ),
                        costs_exchange_rate_currency=(
                            _str_or_none(row[19]) if len(row) > 19 else None
                        ),
                    )
                )
            except Exception as e:
                result.skipped.append(("Dividends", f"Row {row_idx}: {e}"))

    def _parse_bookings(self, ws: Worksheet, result: PDTParseResult) -> None:
        for row_idx, row in _iter_data_rows(ws):
            # 0:broker 1:date 2:time 3:action 4:amount 5:currency
            try:
                broker = _str_or_none(row[0])
                tx_date = _to_date(row[1])
                action = _str_or_none(row[3])
                amount = _float_or_none(row[4])
                currency = _str_or_none(row[5])

                if not all([broker, tx_date, action, amount, currency]):
                    result.skipped.append(
                        ("Bookings", f"Row {row_idx}: missing required fields")
                    )
                    continue

                result.bookings.append(
                    PDTBooking(
                        broker=broker,
                        date=tx_date,
                        action=action,
                        amount=amount,
                        currency=currency,
                    )
                )
            except Exception as e:
                result.skipped.append(("Bookings", f"Row {row_idx}: {e}"))


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------

# 3-row header block for the Transactions sheet
_TX_HEADER_ROW1 = [
    "broker",
    "name",
    "type",
    "search",
    "exchange",
    "date",
    "time",
    "transaction-action",
    "transaction-amount",
    "transaction-price",
    "transaction-price-currency",
    "transaction-price-exchange-rate",
    "transaction-price-exchange-rate-currency",
    "transaction-costs",
    "transaction-costs-currency",
    "transaction-costs-exchange-rate",
    "transaction-costs-exchange-rate-currency",
    "transaction-tax",
    "transaction-tax-currency",
    "transaction-tax-exchange-rate",
    "transaction-tax-exchange-rate-currency",
]
_TX_HEADER_ROW2 = [
    None,
    "Effect",
    None,
    None,
    None,
    "Transaction",
    None,
    None,
    None,
    "Transaction price",
    None,
    None,
    None,
    "Transaction cost",
    None,
    None,
    None,
    "Transaction tax",
    None,
    None,
    None,
]
_TX_HEADER_ROW3 = [
    "Broker",
    "Name",
    "Type",
    "Search",
    "Exchange",
    "Date",
    "Time",
    "Action",
    "Amount",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
]

# 3-row header block for the Dividends sheet
_DIV_HEADER_ROW1 = [
    "broker",
    "name",
    "type",
    "search",
    "exchange",
    "date",
    "time",
    "dividend-action",
    "dividend-amount",
    "dividend-amount-currency",
    "dividend-amount-exchange-rate",
    "dividend-amount-exchange-rate-currency",
    "dividend-tax",
    "dividend-tax-currency",
    "dividend-tax-exchange-rate",
    "dividend-tax-exchange-rate-currency",
    "dividend-costs",
    "dividend-costs-currency",
    "dividend-costs-exchange-rate",
    "dividend-costs-exchange-rate-currency",
]
_DIV_HEADER_ROW2 = [
    None,
    "Effect",
    None,
    None,
    None,
    "Dividend",
    None,
    None,
    "Dividend amount",
    None,
    None,
    None,
    "Dividend tax",
    None,
    None,
    None,
    "Dividend cost",
    None,
    None,
    None,
]
_DIV_HEADER_ROW3 = [
    "Broker",
    "Name",
    "Type",
    "Search",
    "Exchange",
    "Date",
    "Time",
    "Action",
    "Amount",
    "Currency from",
    "Exchange rate",
    "Currency to",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
]

# 3-row header block for the Bookings sheet
_BOOK_HEADER_ROW1 = [
    "broker",
    "date",
    "time",
    "booking-action",
    "booking-amount",
    "booking-amount-currency",
]
_BOOK_HEADER_ROW2 = [None, "Booking", None, None, "Booking amount", None]
# PDT template uses a single space for the broker display label in Bookings
_BOOK_HEADER_ROW3 = [" ", "Date", "Time", "Action", "Value", "Currency"]

# 3-row header block for the Expenses sheet
_EXP_HEADER_ROW1 = [
    "broker",
    "date",
    "time",
    "description",
    "expense-amount",
    "expense-amount-currency",
    "expense-amount-exchange-rate",
    "expense-amount-exchange-rate-currency",
]
_EXP_HEADER_ROW2 = [None, "Expense", None, None, "Expense amount", None, None, None]
_EXP_HEADER_ROW3 = [
    "Broker",
    "Date",
    "Time",
    "Description",
    "Value",
    "Currency from",
    "Exchange rate",
    "Currency to",
]

# PDT Settings sheet — version and API base URL
_PDT_SETTINGS_VERSION = 2.0
_PDT_SETTINGS_URL = "https://beta.portfoliodividendtracker.com/api/import/data"


_ISIN_EXCHANGE_FALLBACK: dict = {
    "US": "Nasdaq",
    "DE": "XETRA Exchange",
    "GB": "London Stock Exchange",
    "FR": "Euronext",
    "NL": "Euronext",
    "BE": "Euronext",
}


def _pdt_exchange(
    exchange: Optional[str],
    asset_type: str,
    symbol: Optional[str] = None,
) -> str:
    """Return the PDT-canonical exchange value for a given asset type.

    Crypto and commodity assets must have an empty exchange field per PDT spec.
    When exchange is None for a stock/ETF/index with an ISIN symbol, infer a
    reasonable default from the ISIN country prefix to avoid PDT validation
    errors.
    """
    if asset_type.lower() in ("crypto", "commodity"):
        return ""
    if exchange:
        return exchange
    # Try to infer from ISIN country prefix (first 2 chars)
    if symbol and len(symbol) == 12 and symbol[:2].isalpha():
        prefix = symbol[:2].upper()
        inferred = _ISIN_EXCHANGE_FALLBACK.get(prefix)
        if inferred:
            return inferred
    return ""


class PDTXLSXExporter:
    """Exports portfolio data to the PDT v2 XLSX format."""

    def export(
        self,
        db_adapter: Any,
        output_path: str,
        portfolio_id: Optional[int] = None,
    ) -> None:
        """
        Export all transactions and dividends to a PDT-compatible XLSX file.

        Args:
            db_adapter: DatabaseAdapter providing access to transactions and assets.
            output_path: Destination .xlsx file path.
            portfolio_id: If given, export only transactions from this portfolio.
        """
        wb = Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        self._write_transactions_sheet(wb, db_adapter, portfolio_id)
        self._write_dividends_sheet(wb, db_adapter, portfolio_id)
        self._write_bookings_sheet(wb, db_adapter, portfolio_id)
        self._write_expenses_sheet(wb)
        self._write_settings_sheet(wb)

        wb.save(output_path)

    def _get_transactions(self, db_adapter: Any, portfolio_id: Optional[int]) -> list:
        if portfolio_id is not None:
            return db_adapter.get_transactions_by_portfolio(portfolio_id)
        return db_adapter.get_all_transactions()

    def _get_portfolio_name(self, db_adapter: Any, portfolio_id: Optional[int]) -> str:
        if portfolio_id is None:
            return ""
        data = db_adapter.get_portfolio(portfolio_id)
        return data.get("name", "") if data else ""

    def _write_transactions_sheet(
        self, wb: Workbook, db_adapter: Any, portfolio_id: Optional[int]
    ) -> None:
        ws = wb.create_sheet("Transactions")
        ws.append(_TX_HEADER_ROW1)
        ws.append(_TX_HEADER_ROW2)
        ws.append(_TX_HEADER_ROW3)

        all_tx = self._get_transactions(db_adapter, portfolio_id)
        for tx in all_tx:
            tx_type = tx.get("transaction_type", "")
            pdt_action = _tx_type_to_pdt_action(tx_type)
            if pdt_action is None:
                continue  # skip dividend / split / transfer rows

            asset = db_adapter.get_asset(tx["asset_id"])
            if not asset:
                continue

            broker = self._resolve_broker(db_adapter, tx)
            asset_type = asset.get("asset_type", "stock")
            pdt_type = _asset_type_to_pdt_type(asset_type)
            tx_date = tx.get("transaction_date")
            if isinstance(tx_date, str):
                try:
                    tx_date = datetime.strptime(tx_date, "%Y-%m-%d").date()
                except ValueError:
                    pass

            fees = tx.get("fees") or 0
            tax = tx.get("tax") or 0
            # tx["currency"] is COALESCE(t.currency, a.currency) — correct per transaction
            currency = tx.get("currency") or asset.get("currency", "EUR")
            exchange = _pdt_exchange(
                asset.get("exchange"), asset_type, asset.get("symbol")
            )

            ws.append(
                [
                    broker,
                    asset.get("name", ""),
                    pdt_type,
                    asset.get("symbol", ""),
                    exchange,
                    tx_date,
                    time(0, 0),
                    pdt_action,
                    tx.get("quantity"),
                    tx.get("price"),
                    currency,
                    None,
                    None,
                    fees if fees else None,
                    currency if fees else None,
                    None,
                    None,
                    tax if tax else None,
                    currency if tax else None,
                    None,
                    None,
                ]
            )

    def _write_dividends_sheet(
        self, wb: Workbook, db_adapter: Any, portfolio_id: Optional[int]
    ) -> None:
        ws = wb.create_sheet("Dividends")
        ws.append(_DIV_HEADER_ROW1)
        ws.append(_DIV_HEADER_ROW2)
        ws.append(_DIV_HEADER_ROW3)

        all_tx = self._get_transactions(db_adapter, portfolio_id)
        for tx in all_tx:
            if tx.get("transaction_type") != "dividend":
                continue

            asset = db_adapter.get_asset(tx["asset_id"])
            if not asset:
                continue

            broker = self._resolve_broker(db_adapter, tx)
            asset_type = asset.get("asset_type", "stock")
            pdt_type = _asset_type_to_pdt_type(asset_type)
            tx_date = tx.get("transaction_date")
            if isinstance(tx_date, str):
                try:
                    tx_date = datetime.strptime(tx_date, "%Y-%m-%d").date()
                except ValueError:
                    pass

            # tx["currency"] is COALESCE(t.currency, a.currency) — correct per transaction
            currency = tx.get("currency") or asset.get("currency", "EUR")
            # total_amount is the dividend received; price*quantity may also work
            dividend_amount = tx.get("total_amount") or (
                (tx.get("price") or 0) * (tx.get("quantity") or 1)
            )
            tax = tx.get("tax") or 0
            exchange = _pdt_exchange(
                asset.get("exchange"), asset_type, asset.get("symbol")
            )

            ws.append(
                [
                    broker,
                    asset.get("name", ""),
                    pdt_type,
                    asset.get("symbol", ""),
                    exchange,
                    tx_date,
                    time(10, 0),
                    "Cash",
                    dividend_amount,
                    currency,
                    None,
                    None,
                    tax if tax else None,
                    currency if tax else None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )

    def _write_bookings_sheet(
        self, wb: Workbook, db_adapter: Any, portfolio_id: Optional[int]
    ) -> None:
        ws = wb.create_sheet("Bookings")
        ws.append(_BOOK_HEADER_ROW1)
        ws.append(_BOOK_HEADER_ROW2)
        ws.append(_BOOK_HEADER_ROW3)

        try:
            bookings = db_adapter.get_all_bookings(portfolio_id)
        except Exception:
            bookings = []

        for bk in bookings:
            broker = ""
            pid = bk.get("portfolio_id")
            if pid:
                portfolio = db_adapter.get_portfolio(pid)
                if portfolio:
                    broker = portfolio.get("name", "")

            bk_date = bk.get("date")
            if isinstance(bk_date, str):
                try:
                    bk_date = datetime.strptime(bk_date, "%Y-%m-%d").date()
                except ValueError:
                    pass

            ws.append(
                [
                    broker,
                    bk_date,
                    time(9, 0),
                    bk.get("action", "Deposit"),
                    bk.get("amount"),
                    bk.get("currency", "EUR"),
                ]
            )

    def _write_expenses_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Expenses")
        ws.append(_EXP_HEADER_ROW1)
        ws.append(_EXP_HEADER_ROW2)
        ws.append(_EXP_HEADER_ROW3)

    def _write_settings_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Settings")
        ws.append(["Version", "URL"])
        ws.append([_PDT_SETTINGS_VERSION, _PDT_SETTINGS_URL])

    def _resolve_broker(self, db_adapter: Any, tx: dict) -> str:
        """Return portfolio name as broker, or empty string if none."""
        portfolio_id = tx.get("portfolio_id")
        if portfolio_id:
            data = db_adapter.get_portfolio(portfolio_id)
            if data:
                return data.get("name", "")
        return ""


# ---------------------------------------------------------------------------
# Convenience functions
# ---------------------------------------------------------------------------


def parse_pdt_xlsx(file_path: str) -> PDTParseResult:
    """Parse a PDT XLSX file and return structured data."""
    return PDTXLSXParser().parse(file_path)


def export_pdt_xlsx(
    db_adapter: Any,
    output_path: str,
    portfolio_id: Optional[int] = None,
) -> None:
    """Export portfolio data to a PDT-compatible XLSX file."""
    PDTXLSXExporter().export(db_adapter, output_path, portfolio_id)
